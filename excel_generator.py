import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os
import zipfile
from config import TEMP_DIR

def create_reviews_excel(reviews, article, filter_label):
    """Создаёт Excel с отзывами."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отзывы"

    headers = ["Дата", "Автор", "Оценка", "Текст", "Достоинства", "Недостатки",
               "Ответ продавца", "Дата ответа", "Фото/Видео"]

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6B2FA0", end_color="6B2FA0", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, review in enumerate(reviews, 2):
        values = [review["date"], review["author"], review["rating"], review["text"],
                  review["pros"], review["cons"], review["seller_reply"],
                  review["reply_date"], review["media_count"]]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    # Ширина столбцов
    widths = [12, 18, 8, 50, 30, 30, 30, 12, 15]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Автофильтр
    ws.auto_filter.ref = ws.dimensions

    filename = os.path.join(TEMP_DIR, f"Отзывы_{article}_{filter_label}.xlsx")
    wb.save(filename)
    return filename

def create_questions_excel(questions, article):
    """Создаёт Excel с вопросами."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Вопросы"

    headers = ["Дата", "Автор", "Текст вопроса", "Ответ продавца", "Дата ответа", "Ответы других покупателей"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="E05B2C", end_color="E05B2C", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    for row_idx, q in enumerate(questions, 2):
        values = [q["date"], q["author"], q["text"], q["seller_reply"],
                  q["reply_date"], q["other_answers"]]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    widths = [12, 18, 50, 40, 12, 40]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.auto_filter.ref = ws.dimensions

    filename = os.path.join(TEMP_DIR, f"Вопросы_{article}.xlsx")
    wb.save(filename)
    return filename

def create_archive(reviews_file, questions_file, article, filter_label):
    """Создаёт ZIP-архив с обоими файлами."""
    archive_name = os.path.join(TEMP_DIR, f"Архив_{article}_{filter_label}.zip")
    with zipfile.ZipFile(archive_name, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.write(reviews_file, os.path.basename(reviews_file))
        zf.write(questions_file, os.path.basename(questions_file))
    return archive_name